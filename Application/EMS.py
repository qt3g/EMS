import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import os

user = os.getenv('USERNAME')
output_dir = ""
default_output_dir = f"C:/Users/{user}/Desktop"
file_name_format = "{month}月_{year}.xlsx"

def convert_to_reiwa(year):
    return year - 2018

def convert_to_seireki(year):
    return year + 2018

def split_excel_by_month(file_path, reiwa_year):
    seireki_year = convert_to_seireki(reiwa_year)
    df = pd.read_excel(file_path, engine='openpyxl')

    df['月'] = df['月'].fillna(0).astype(int).astype(str)
    df['日'] = df['日'].fillna(0).astype(int)

    for month, group in df.groupby('月'):
        group = group.sort_values(by='日')
        output_file = os.path.join(output_dir, file_name_format.format(month=month, year=seireki_year))
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            income_sum = group['収入'].sum()
            expense_sum = group['支出'].sum()
            balance = income_sum - expense_sum
            
            total_row = pd.DataFrame({'月': ['合計'], '収入': [income_sum], '支出': [expense_sum], '備考': [balance]})
            group = pd.concat([group, total_row], ignore_index=True)
            
            group.to_excel(writer, index=False, startrow=1)
            worksheet = writer.sheets['Sheet1']
            worksheet['A1'] = f'令和{reiwa_year}年【{month}月】({seireki_year}年)'
            worksheet['A1'].font = Font(bold=True)
            
            fill = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
            font_white_bold = Font(color="FFFFFF", bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col in range(1, 8):
                cell = worksheet.cell(row=2, column=col)
                cell.fill = fill
                cell.font = font_white_bold
                cell.border = thin_border
            
            for row in range(3, group.shape[0] + 3):
                for col in range(1, 8):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                date_cell = worksheet.cell(row=row, column=2)
                date_cell.number_format = 'd'
            
            balance_cell = worksheet.cell(row=group.shape[0] + 2, column=7)
            balance_cell.value = balance
            balance_cell.font = Font(bold=True, color="FF0000")
            balance_cell.border = thin_border
            
            worksheet.column_dimensions['A'].width = 3.00
            worksheet.column_dimensions['B'].width = 3.00
            worksheet.column_dimensions['C'].width = 22.00
            worksheet.column_dimensions['D'].width = 30.00
            worksheet.column_dimensions['E'].width = 10.00
            worksheet.column_dimensions['F'].width = 10.00
            worksheet.column_dimensions['G'].width = 10.00

def on_button_click():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            reiwa_year = int(reiwa_entry.get())
            split_excel_by_month(file_path, reiwa_year)
            current_time = datetime.now().strftime("%Y/%m/%d %H:%M")
            log_message = f"各月のファイルが生成されました {current_time}"
            log_textbox.insert(tk.END, log_message + "\n", "success")
            log_textbox.see(tk.END)
            print(log_message)
    except Exception as e:
        error_message = f"エラーが発生しました: {str(e)}"
        log_textbox.insert(tk.END, error_message + "\n", "error")
        log_textbox.see(tk.END)
        print(error_message)
        with open("error_log.txt", "a") as error_log:
            error_log.write(f"{datetime.now()}: {str(e)}\n")

def show_tutorial():
    tutorial_message = (
        "使い方:\n"
        "1. 令和の年を入力してください。\n"
        "   - 例: 令和3年の場合は「3」と入力します。\n"
        "2. 「メニュー」から「ディレクトリを指定」を選択し、出力先ディレクトリを設定してください。\n"
        "   - ここで指定したフォルダに分割されたファイルが保存されます。\n"
        "3. 「Excelファイルを選択して月ごとに分割」ボタンをクリックしてください。\n"
        "   - 分割したいExcelファイルを選択します。\n"
        "4. ファイルを選択すると、各月ごとに分割されたファイルが生成されます。\n"
        "   - 生成されたファイルは指定した出力先ディレクトリに保存されます。\n"
        "5. 「ログをクリア」ボタンでログをクリアできます。\n"
        "   - 操作履歴やエラーメッセージをクリアします。\n"
        "6. 「メニュー」から「ファイル名フォーマットを設定」を選択し、出力ファイルの名前フォーマットを指定できます。\n"
        "   - 例: 「{month}月_{year}.xlsx」と入力すると、「3月_2021.xlsx」のようなファイル名になります。\n"
    )
    messagebox.showinfo("使い方", tutorial_message)
        
def set_output_directory():
    global output_dir
    output_dir = filedialog.askdirectory()
    if output_dir:
        log_message = f"出力ディレクトリが設定されました: {output_dir}"
        log_textbox.insert(tk.END, log_message + "\n", "success")
        log_textbox.see(tk.END)
        print(log_message)

def clear_log():
    log_textbox.delete(1.0, tk.END)

def set_file_name_format():
    global file_name_format
    file_name_format = file_name_format_entry.get()
    log_message = f"ファイル名フォーマットが設定されました: {file_name_format}"
    log_textbox.insert(tk.END, log_message + "\n", "success")
    log_textbox.see(tk.END)
    print(log_message)

def show_file_name_format_dialog():
    global file_name_format_entry
    dialog = tk.Toplevel(root)
    dialog.title("ファイル名フォーマットの設定")
    dialog.geometry("400x200")
    
    tk.Label(dialog, text="ファイル名フォーマット:", font=("Arial", 12)).pack(pady=10)
    file_name_format_entry = tk.Entry(dialog, font=("Arial", 12))
    file_name_format_entry.pack(pady=10)
    file_name_format_entry.insert(0, file_name_format)
    
    set_format_button = tk.Button(dialog, text="フォーマットを設定", command=set_file_name_format, font=("Arial", 12))
    set_format_button.pack(pady=10)

root = tk.Tk()
root.title("Excel Splitter")
root.geometry("500x500")

menu = tk.Menu(root)
root.config(menu=menu)
file_menu = tk.Menu(menu)
menu.add_cascade(label="メニュー", menu=file_menu)
file_menu.add_command(label="ディレクトリを指定", command=set_output_directory)
file_menu.add_command(label="ファイル名フォーマットを設定", command=show_file_name_format_dialog)

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(pady=20)

tk.Label(frame, text="令和何年:", font=("Arial", 12)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
reiwa_entry = tk.Entry(frame, font=("Arial", 12))
reiwa_entry.grid(row=0, column=1, padx=5, pady=5)

button = tk.Button(frame, text="Excelファイルを選択して月ごとに分割", command=on_button_click, font=("Arial", 12))
button.grid(row=1, column=0, columnspan=2, pady=10)

tutorial_button = tk.Button(frame, text="使い方", command=show_tutorial, font=("Arial", 12))
tutorial_button.grid(row=2, column=0, columnspan=2, pady=10)

clear_log_button = tk.Button(frame, text="ログをクリア", command=clear_log, font=("Arial", 12))
clear_log_button.grid(row=3, column=0, columnspan=2, pady=10)

log_textbox = tk.Text(root, height=8, wrap=tk.WORD, font=("Arial", 10))
log_textbox.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

log_textbox.tag_configure("error", foreground="red")
log_textbox.tag_configure("success", foreground="green")

output_dir = default_output_dir
log_message = f"デフォルトの出力ディレクトリが設定されました: {output_dir}"
log_textbox.insert(tk.END, log_message + "\n", "success")
log_textbox.see(tk.END)
print(log_message)

root.mainloop()